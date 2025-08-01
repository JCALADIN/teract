from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List


def ensure_columns(ws: Worksheet, header_row: int, groups_row: int, ia_cols: List[str]) -> Dict[str, int]:
    """
    Ajoute les colonnes IA manquantes à droite et renvoie le mapping
    {col_name: column_index}.
    """
    existing = {cell.value: idx for idx, cell in enumerate(ws[header_row], 1)}
    cur_max  = ws.max_column

    for col in ia_cols:
        if col not in existing:
            cur_max += 1
            ws.cell(row=header_row, column=cur_max, value=col)      # ligne 2
            ws.cell(row=groups_row, column=cur_max, value="")       # ligne 1 vide
            existing[col] = cur_max
    return existing
