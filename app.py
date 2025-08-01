# app.py ────────────────────────────────────────────────────────────────
# Streamlit • Générateur Marketing IA – Teract Corporate Edition
# -----------------------------------------------------------------------

import sys, importlib.util, re
from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook

from style import apply_style
import settings
from columns import BASE_COLUMNS, IA_COLUMNS
from data_io import read_file
from st_utils import preview_df
from llm_utils import build_user_prompt, call_llm
from docs import GUIDE_MD
from xl_utils import ensure_columns
# import notify   # notification e-mail désactivée pour l’instant

# ───────────────────────────────
# 0. Style global
# ───────────────────────────────
apply_style()

# ───────────────────────────────
# 1. Sidebar : guide utilisateur
# ───────────────────────────────
if "show_doc" not in st.session_state:
    st.session_state["show_doc"] = False

def _toggle_doc():
    st.session_state["show_doc"] = not st.session_state["show_doc"]

st.sidebar.button("📖 Guide utilisateur", on_click=_toggle_doc)
if st.session_state["show_doc"]:
    st.sidebar.markdown(GUIDE_MD, unsafe_allow_html=True)

# ───────────────────────────────
# 2. Upload fichier
# ───────────────────────────────
st.markdown("### Chargez votre fichier Teract")
uploaded = st.file_uploader(
    "Formats acceptés : xls, xlsm, xlsx, csv",
    type=["xls", "xlsm", "xlsx", "csv"],
)

if uploaded:
    try:
        # Lecture générique
        df, wb, ws_entities, groups_row = read_file(uploaded)

        # Normalisation des en-têtes
        df.columns = [
            re.sub(r"\s+", " ", c.replace("/", " ")).strip() for c in df.columns
        ]

        # Validation colonnes obligatoires
        missing = [c for c in BASE_COLUMNS if c not in df.columns]
        if missing:
            st.markdown(
                f"<div class='error-msg'>❌ Colonnes manquantes : {', '.join(missing)}</div>",
                unsafe_allow_html=True,
            )
            st.stop()

        # Ajout des colonnes IA manquantes (DataFrame)
        for col in IA_COLUMNS:
            if col not in df.columns:
                df[col] = ""

        st.markdown(
            "<div class='success-msg'>✅ Fichier conforme.</div>",
            unsafe_allow_html=True,
        )
        preview_df(df, "Aperçu de la feuille Entities", "input_preview")

    except Exception as err:
        st.error(f"Erreur : {err}")
        st.stop()

    # ───────────────────────────────
    # 3. Génération IA (boucle corrigée)
    # ───────────────────────────────
    if st.button("🚀 Lancer la génération IA"):
        total        = len(df)
        errors       = 0
        total_tokens = 0

        bar = st.progress(0.0)
        txt = st.empty()

        # Préparer le mapping colonnes → index pour openpyxl (si classeur multi-onglets)
        if wb:
            col_map = ensure_columns(
                ws_entities,
                header_row=2,   # ligne 2 = noms de colonnes
                groups_row=1,   # ligne 1 = groupes
                ia_cols=IA_COLUMNS,
            )

        for idx, row in df.iterrows():
            try:
                rep = call_llm(build_user_prompt(row))
                # --- DataFrame ---
                df.at[idx, "Description Marketing Client 1"] = rep["desc"]
                df.at[idx, "Plus produit 1"]                = rep["plus1"]
                df.at[idx, "Plus produit 2"]                = rep["plus2"]
                df.at[idx, "Plus produit 3"]                = rep["plus3"]
                df.at[idx, "IA DATA"]  = 1
                df.at[idx, "IAPLUS"]   = 1
                df.at[idx, "Token"]    = rep["tokens"]
                df.at[idx, "Date"]     = datetime.now().strftime("%d/%m/%Y %H:%M")
                df.at[idx, "Commentaires"] = ""
                total_tokens += rep["tokens"]

                # --- Écriture directe dans Excel ---
                if wb:
                    xl_row = idx + 3  # données commencent ligne 3
                    ws_entities.cell(xl_row, col_map["Description Marketing Client 1"], rep["desc"])
                    ws_entities.cell(xl_row, col_map["Plus produit 1"],                rep["plus1"])
                    ws_entities.cell(xl_row, col_map["Plus produit 2"],                rep["plus2"])
                    ws_entities.cell(xl_row, col_map["Plus produit 3"],                rep["plus3"])
                    ws_entities.cell(xl_row, col_map["IA DATA"],  1)
                    ws_entities.cell(xl_row, col_map["IAPLUS"],   1)
                    ws_entities.cell(xl_row, col_map["Token"],    rep["tokens"])
                    ws_entities.cell(xl_row, col_map["Date"],     datetime.now().strftime("%d/%m/%Y %H:%M"))
                    ws_entities.cell(xl_row, col_map["Commentaires"], "")

            except Exception as e:
                errors += 1
                df.at[idx, "IA DATA"] = 0
                df.at[idx, "IAPLUS"]  = 0
                df.at[idx, "Commentaires"] = str(e)[:250]

                if wb:
                    xl_row = idx + 3
                    ws_entities.cell(xl_row, col_map["IA DATA"], 0)
                    ws_entities.cell(xl_row, col_map["IAPLUS"], 0)
                    ws_entities.cell(xl_row, col_map["Commentaires"], str(e)[:250])

            # Mise à jour barre et texte
            bar.progress((idx + 1) / total)
            txt.markdown(
                f"<div style='text-align:center;font-weight:600;color:var(--primary);'>"
                f"{idx + 1}/{total} lignes traitées</div>",
                unsafe_allow_html=True,
            )

        bar.empty(); txt.empty()
        st.success(f"Génération terminée : {total - errors} lignes OK, {errors} erreurs.")

        preview_df(df, "Aperçu de la feuille enrichie", "output_preview")

        # ───────────────────────────────
        # 4. Export
        # ───────────────────────────────
        buf = BytesIO()
        if wb:   # classeur multi-feuilles (on garde macros et formats)
            wb.save(buf)
            ext  = ".xlsm" if uploaded.name.lower().endswith(".xlsm") else ".xlsx"
            mime = (
                "application/vnd.ms-excel"
                if ext == ".xlsm"
                else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:    # CSV ou Excel simple
            has_xlwt = importlib.util.find_spec("xlwt") and sys.version_info < (3, 12)
            eng, ext, mime = (
                ("xlwt", ".xls", "application/vnd.ms-excel")
                if has_xlwt
                else (
                    "openpyxl",
                    ".xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            )
            df.to_excel(buf, index=False, engine=eng)

        buf.seek(0)
        st.download_button(
            "💾 Télécharger le fichier",
            buf,
            file_name=f"catalogue_enrichi_{datetime.now():%Y%m%d_%H%M}{ext}",
            mime=mime,
        )

        # notify.send_report(...)  # e-mail inactif

else:
    st.info("Déposez un fichier pour commencer.")

st.markdown(
    "---\n<div style='text-align:center;font-size:0.85rem;'>© 2025 Teract – "
    "Réservé au service Innovation.</div>",
    unsafe_allow_html=True,
)
