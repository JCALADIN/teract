from typing import Tuple, Optional
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def _read_csv(upload) -> Tuple[pd.DataFrame, None, None, None]:
    try:
        df = pd.read_csv(upload, dtype=str).fillna("")
    except pd.errors.ParserError:
        df = pd.read_csv(upload, sep=";", dtype=str).fillna("")
    return df, None, None, None


def _read_excel_single(upload, engine):
    df = pd.read_excel(upload, dtype=str, engine=engine).fillna("")
    return df, None, None, None


def _read_excel_multi(upload) -> Tuple[pd.DataFrame, Workbook, Worksheet, list]:
    wb = load_workbook(upload, keep_vba=True, data_only=False)
    if "Entities" not in wb.sheetnames:
        raise ValueError("Classeur : la feuille « Entities » est absente.")
    ws: Worksheet = wb["Entities"]

    if ws.max_row < 2:
        raise ValueError("Feuille Entities : moins de deux lignes (groupes + en-têtes).")

    groups_row = [cell.value or "" for cell in ws[1]]
    headers    = [cell.value or "" for cell in ws[2]]

    # DataFrame des lignes 3 → N
    data = [[cell.value or "" for cell in row] for row in ws.iter_rows(min_row=3, max_row=ws.max_row)]
    df = pd.DataFrame(data, columns=headers).fillna("")

    return df, wb, ws, groups_row


def read_file(upload):
    """
    Retourne:
        df          → DataFrame à traiter
        workbook    → Workbook openpyxl  (ou None)
        ws_entities → Worksheet          (ou None)
        groups_row  → list (ligne1)      (ou None)
    """
    name = upload.name.lower()
    if name.endswith(".csv"):
        return _read_csv(upload)

    if name.endswith((".xls", ".xlsm", ".xlsx")):
        engine = "openpyxl" if name.endswith(".xlsx") else None
        with pd.ExcelFile(upload, engine=engine) as xf:
            if len(xf.sheet_names) == 1:
                return _read_excel_single(upload, engine)
        # Plusieurs feuilles  → openpyxl
        upload.seek(0)  # reset pointer
        return _read_excel_multi(BytesIO(upload.read()))

    raise ValueError("Format non pris en charge (CSV, XLS, XLSX, XLSM).")
